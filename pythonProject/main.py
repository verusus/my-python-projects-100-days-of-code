from configparser import ConfigParser
import openpyxl
from openpyxl.comments import Comment
import oracledb
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import column_index_from_string
import glob

configparser = ConfigParser()
configparser.read("config.ini")

LIMIT_HEADER_SEARCH = 50
SCHEMA = configparser['database']['SCHEMA']
PATH = configparser['directory_path']['PATH']


# connect to database
connection = oracledb.connect(user=configparser['database']['USER'],
                              password=configparser['database']['PASSWORD'],
                              dsn=configparser['database']['DSN'])
cursor = connection.cursor()


# -----------------------------------------FIND THE HEADER_ROW_INDEX --------------------------------
def get_header_row_index(id_doc, worksheet: openpyxl.workbook.workbook.Worksheet):
    query_ = f"SELECT c_colonne, id_col from {SCHEMA}.COREP_COL WHERE id_doc='{id_doc}' AND c_colonne IS NOT NULL"
    first_record = cursor.execute(query_).fetchone()

    # col_name, cell_content_db = first_record[3].strip("C_COLONNE_"), first_record[2]
    try:
        col_name_, cell_content_db = first_record[0][10:], first_record[1]
        col_index_ = column_index_from_string(col_name_)
    except:
        return None

    # searching for the cells' header_row_index
    cell_not_found = True
    header_row_index = 1
    while cell_not_found:
        cell_val = worksheet.cell(row=header_row_index, column=col_index_).value
        if cell_val == cell_content_db:
            cell_not_found = False
        elif header_row_index > LIMIT_HEADER_SEARCH:
            return None
        else:
            header_row_index += 1
    return header_row_index


# -----------------------------------------FIND THE HEADER_COLUMN_INDEX --------------------------------
def get_header_column_index(id_doc, worksheet: openpyxl.workbook.workbook.Worksheet):
    query_ = f"SELECT no_excel, id_ligne from {SCHEMA}.COREP_LIGNE WHERE ID_DOC='{id_doc}' AND no_excel IS NOT NULL" \
             f" ORDER BY id_ligne ASC"
    record = cursor.execute(query_).fetchone()

    try:
        row_index, cell_content_db = record[0], record[1]
        # searching for the cells' header_row_index
        cell_not_found = True
        header_column_index = 1
        while cell_not_found:
            cell_val = worksheet.cell(column=header_column_index, row=row_index).value
            if cell_val == cell_content_db:
                cell_not_found = False
            elif header_column_index > LIMIT_HEADER_SEARCH:
                return None
            else:
                header_column_index += 1
        return header_column_index
    except:
        return None


# ---------------------------------------- ADD COMMENTS FOR ROW HEADER --------------------------------------
def get_ss_docs_map(id_doc):
    """This method returns a map of comments by id_ss_doc as dict if there is a ss_doc for the current doc given by id_doc.
    comments like {"a": "FROM 010 TO 030:",...}. else {None: "FOR ALL ID_LIGNES:"}"""
    query = f"SELECT * FROM {SCHEMA}.corep_ligne WHERE ID_DOC='{id_doc}' AND id_ss_doc IS NOT NULL ORDER BY id_ligne ASC"
    records = cursor.execute(query).fetchall()
    doc_map = {}
    i = 0
    if len(records) > 0:
        while i < len(records):
            id_ss_doc = records[i][9]
            id_ligne_start_pointer = records[i][2]
            while i < len(records) and id_ss_doc == records[i][9]:
                id_ligne_end_pointer = records[i][2]
                i += 1
            if id_ligne_start_pointer == id_ligne_end_pointer:
                doc_map[id_ss_doc] = f"FOR ID_LIGNE {id_ligne_start_pointer}:\n"
            else:
                doc_map[id_ss_doc] = f"FROM ID_LIGNE {id_ligne_start_pointer} TO {id_ligne_end_pointer}:\n"
        return doc_map
    else:
        doc_map[None] = "FOR ALL ID_LIGNES:\n"
        return doc_map


def fill_header_row(id_doc, worksheet: openpyxl.workbook.workbook.Worksheet):
    HEADER_ROW_INDEX = get_header_row_index(id_doc, worksheet)
    # CHECK FOR THE HEADER EXISTANCE
    if HEADER_ROW_INDEX is not None:
        # CHECK IF THE DOC HAS SS_DOCS
        ss_docs_map = get_ss_docs_map(id_doc)
        query = f"SELECT * FROM {SCHEMA}.corep_col WHERE id_doc='{id_doc}' AND c_colonne IS NOT NULL ORDER BY id_col, id_ss_doc ASC"
        records = cursor.execute(query).fetchall()

        if len(records) > 0:
            i = 0
            while i < len(records):
                cell_content = records[i][2]  # id_col = cell_content
                formated_comment = ""
                col_name = records[i][3][10:]
                col_index = column_index_from_string(col_name)

                while i < len(records) and cell_content == records[i][2]:  # cell content not changed
                    # format comment
                    formated_comment += ss_docs_map[records[i][10]]     # records[i][10] = id_ss_doc

                    formated_comment += f"C_INITIAL_FINAL: {records[i][8]}\nF_NEGATIF: {records[i][9]}\nC_OPERATION: {records[i][6]}\nL_DATA: {records[i][5]}" \
                                        f"\nL_LIB: {records[i][7]}\n\n"
                    i += 1
                comment = Comment(text=formated_comment, author='', height=900, width=800)
                try:
                    worksheet.cell(row=HEADER_ROW_INDEX, column=col_index).comment = comment
                except:
                    print(f"there was an error in cell with cordinates: (row={HEADER_ROW_INDEX}, column={col_index}")


# ---------------------------------------- ADD COMMENTS FOR COLUMN HEADER --------------------------------------
def fill_header_column(id_doc, worksheet: openpyxl.workbook.workbook.Worksheet):
    HEADER_COLUMN_INDEX = get_header_column_index(id_doc, worksheet)
    if HEADER_COLUMN_INDEX is not None:
        query = f"SELECT * from {SCHEMA}.COREP_LIGNE WHERE ID_DOC='{id_doc}' AND no_excel IS NOT NULL ORDER BY id_ligne ASC"
        records = cursor.execute(query).fetchall()
        for record in records:
            formated_comment = f"L_NOM: {record[5]}\nL_ITS: {record[4]}\nC_WHERE_INIT: {record[6]}\nC_WHERE_FINAL: {record[7]}" \
                               f"\nID_SS_DOC: {record[9]}"
            comment = Comment(text=formated_comment, author='', height=400, width=400)
            try:
                worksheet.cell(row=record[8], column=HEADER_COLUMN_INDEX).comment = comment
            except:
                print(f"there was an error in cell with cordinates: (row={record[8]}, column={HEADER_COLUMN_INDEX}")


# ---------------------------------------- ADD COMMENTS FROM COREP_DIM --------------------------------------
def fill_corep_dim_comments(id_doc, worksheet: openpyxl.workbook.workbook.Worksheet):
    query = f"SELECT * from {SCHEMA}.corep_dim WHERE id_doc='{id_doc}' ORDER BY id_dim ASC"
    records = cursor.execute(query).fetchall()
    column_i = 1
    for record in records:
        formated_comment = f"ID_DIM: {record[2]}\nL_ITS: {record[6]}\nC_WHERE_INIT: {record[8]}\nC_WHERE_FINAL: {record[9]}" \
                           f"\nL_NOM: {record[7]}"
        comment = Comment(text=formated_comment, author='', height=400, width=600)
        try:
            worksheet.cell(column=column_i, row=1).comment = comment
        except:
            print(f"there was an error affecting comment to cell(column={column_i}, row=1")
            column_i += 1
        column_i += 1


# ---------------------------------------- ADD COMMENTS FROM COREP_DOC FOR EACH FILE -------------------------
files_paths = glob.glob(PATH + "/*.xlsx")

for file_path in files_paths:

    wb = openpyxl.load_workbook(file_path)
    query = f"SELECT * FROM {SCHEMA}.corep_doc WHERE n_doc IS NOT NULL ORDER BY n_doc ASC"
    records = cursor.execute(query).fetchall()
    for record in records:
        if record[10] in wb.sheetnames:
            formated_comment = f"F_ACTIF: {record[5]}\nL_TABLE: {record[8]}\nC_WHERE_INIT: {record[6]}\nC_WHERE_FINAL: {record[7]}"
            comment = Comment(text=formated_comment, author='', height=400, width=600)
            try:
                wb[record[10]].cell(row=2, column=1).comment = comment
            except:
                print(f"There was a problem with cell(row=2, column=1) in sheet {record[10]}")
            # while in each sheet, fill all the comments from corep_col, corep_ligne, corep_dim
            # check if there is a header first
            fill_header_row(id_doc=record[1], worksheet=wb[record[10]])
            fill_header_column(id_doc=record[1], worksheet=wb[record[10]])
            fill_corep_dim_comments(id_doc=record[1], worksheet=wb[record[10]])

            # add colors
            wb[record[10]].sheet_properties.tabColor = "03C988"
            row_i = 1
            row_i_not_found = True
            while row_i_not_found and row_i < 200:
                cell_val = wb["Index"].cell(row=row_i, column=3).value
                if cell_val == record[2]:
                    row_i_not_found = False
                    wb["Index"].cell(row=row_i, column=3).fill = PatternFill(start_color="03C988", end_color="03C988",
                                            fill_type = "solid")
                row_i += 1
    # SAVE AND CLOSE THE FILE
    wb.save(file_path)
    wb.close()

# -------------------------------------- CLOSE RESOURCES --------------------------------------
cursor.close()
connection.close()
