from openpyxl import load_workbook

wb = load_workbook(filename = '/Users/YOUSSRA/OneDrive/Bureau/excel/Fichetelecoms.xlsx', read_only = 'True')
sheetDR = wb['DR Doukkala']
sheetFM = wb['FesMeknes ']
sheetOr = wb['Orientale ']
sheetElhaouz = wb['Elhaouz']
sheetTad = wb['Tadla']
sheetSud = wb['Sud']
sheetRT = wb['Rabat-Temara']
sheetSale = wb['Sale ELgharb']
sheetCasaNord = wb['Casa Nord']
sheetCasaSud = wb['casa Sud ']
sheetNord = wb['Nord']
sheetmobile = wb['flotte mobile']
num=input("entrer: ")
def create_table(sheet):
    List = []
    for value in sheet.iter_rows(min_row=2, max_row=200, min_col=2, max_col=8, values_only=True):
        List.append(value)
    return List

def find_from_num(num, List):
    for i in range (0, len(List)):
        if List[i][2] == num:
            return i
    return False
def parcours_ws_num(num, sheet):
    L = create_table(sheet)
    r = find_from_num(num, L)
    print('r = %int', r)
    print (L[r])
    return r
parcours_ws_num(num,sheetmobile )

