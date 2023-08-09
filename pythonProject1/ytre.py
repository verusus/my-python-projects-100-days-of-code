import mysql.connector
import tkinter
import time
from openpyxl import load_workbook
import tkinter as tk


database=mysql.connector.connect(
host="localhost",
username= "root",
password="fban1937",
database="admin",
)
num = ""

wb = load_workbook(filename = '/Users/YOUSSRA/OneDrive/Bureau/excel/Fichetelecoms.xlsx', read_only = 'True')
sheetDR = wb['DR Doukkala']
sheetFM = wb['FesMeknes ']
sheetOr = wb['Orientale ']
sheetElhaouz = wb['Elhaouz']
sheetTad = wb['Tadla']
sheetSud = wb['Sud']
sheetRT = wb['Rabat-Temara']
sheetSale = wb['Sale ELgharb']
sheetCasaNord = wb['Casa Nord']
sheetCasaSud = wb['casa Sud ']
sheetNord = wb['Nord']
sheetmobile = wb['flotte mobile']

def create():

  curadmin=database.cursor()
  curadmin.execute('SELECT * FROM admin WHERE adminename = %s AND passwordname = %s', (username, password))
  admin= curadmin.fetchone()
  if admin:
        import tkinter as tk
        acceuil=tk.Toplevel()
        acceuil.title(" PAGE D'ACCEUIL")
        acceuil.geometry("270x210")
        acceuil.configure(bg="#39d972")
        registre2=tkinter.Button(acceuil,text="gestion",command=creategestion, bg="#93ff00")
        registre2.pack()
        registre1=tkinter.Button(acceuil,text="enregistrer",command=createenregi, bg="#93ff00")
        registre1.pack()
        registre3=tkinter.Button(acceuil,text="supprimer",command=createsuppression, bg="#93ff00")
        registre3.pack()
        acceuil.mainloop()
        createenregi()
        createsuppression()
        creategestion()
def login():
 # connect
 global username
 global password
 username = user.get()
 password = passw.get()
# Check if account exists using MySQL
 curutilisateur=database.cursor()
 curutilisateur.execute('SELECT * FROM compte WHERE username = %s AND password = %s', (username, password))
 utlisateur= curutilisateur.fetchone()
# If account exists in accounts table in out database
 if not utlisateur:
    messagelogin['text']="erreur"
    return False
 if username and password in utlisateur:
    messagelogin['text']="bienvenue"
    create()

def createenregi():
    app=tk.Toplevel()
    app.title(" PAGE D'ENREGISTREMENT")
#change size
    app.geometry("270x210")
#change window colour
    app.configure(bg="#39d972")
    title2 = tkinter.Label(app, text="_____Enregistrement___", bg="#39d972")
    usertitle2 = tkinter.Label(app, text="---Username---", bg="#39d972")
    user2 = tkinter.Entry(app)
    passtitle2 = tkinter.Label(app, text="---Password---", bg="#39d972")
    passw2= tkinter.Entry(app)
    messageEnregistrement = tkinter.Label(app,text=' ', bg="#39d972")
    regiButton=tkinter.Button(app,text="creer un compte",command=enregistrer, bg="#93ff00")
    title2.pack()
    usertitle2.pack()
    user2.pack()
    passtitle2.pack()
    passw2.pack()
    regiButton.pack()
    messageEnregistrement.pack()
    app.mainloop()
def enregistrer():
    createenregi()
    cur = database.cursor(buffered=True)
    username = user2.get()
    password= passw2.get()
#Check if account exists using MySQL
    cur.execute('SELECT * FROM compte WHERE username = %s',[username])
    enregistrement = cur.fetchone()
        # If account exists show error and validation checks
    if not enregistrement:
        cur.execute('INSERT INTO compte VALUES (%s, %s)', [username, password])
        database.commit()
        messageEnregistrement['text']="enregistrement effecué"
    if enregistrement :
        messageEnregistrement['text']="compte existant"
def createsuppression():
    sup=tk.Toplevel()
    sup.title(" PAGE DE SUPPRESSION")
    sup.geometry("270x210")
    sup.configure(bg="#39d972")
    title3 = tkinter.Label(sup, text="_____Suppression___", bg="#39d972")
    usertitle3 = tkinter.Label( sup, text="---Username---", bg="#39d972")
    user3= tkinter.Entry( sup)
    passtitle3 = tkinter.Label( sup, text="---Password---", bg="#39d972")
    passw3= tkinter.Entry( sup)
    messageSuppression = tkinter.Label( sup,text=' ', bg="#39d972")
    supButton=tkinter.Button(sup,text="supprimer le compte",command=supprimer, bg="#93ff00")
    title3.pack()
    usertitle3.pack()
    user3.pack()
    passtitle3.pack()
    passw3.pack()
    supButton.pack()
    messageSuppression.pack()
    sup.mainloop()
def supprimer():
    createsuppression()
    cursup = database.cursor()
    username = user3.get()
    password= passw3.get()
#Check if account exists using MySQL
    cursup.execute('SELECT * FROM compte WHERE username = %s',[username])
    suprim = cursup.fetchone()
        # If account exists show error and validation checks
    if not suprim:
        messageSuppression['text']="compte inexistant"
    if suprim :
        cursup.execute('DELETE FROM compte WHERE username = %s',[username])
        database.commit()
        messageSuppression['text']="suppression effectuée"
def creategestion():
    ges=tk.Toplevel()
    ges.title(" PAGE DE GESTION")
    ges.geometry("270x210")
    ges.configure(bg="#39d972")
    title4 = tkinter.Label(ges, text="_____Gestion___", bg="#39d972")
    ligne = tkinter.Label( ges, text="N° de ligne", bg="#39d972")
    ligneentre= tkinter.Entry( ges)
    forfait = tkinter.Label( ges, text="Forfait", bg="#39d972")
    forfaitentre= tkinter.Entry( ges)
    typeaces = tkinter.Label( ges, text="Type d'accés'", bg="#39d972")
    typeentre= tkinter.Entry( ges)
    facturation = tkinter.Label( ges, text="Facturation", bg="#39d972")
    facturationentre= tkinter.Entry( ges)
    messagegestion = tkinter.Label(ges,text=' ', bg="#39d972")
    gestion=tkinter.Button(ges,text="chercher",command=parcours_ws_num  , bg="#93ff00")
    title4.pack( )
    ligne.pack()
    ligneentre.pack()
    forfait.pack()
    forfaitentre.pack( )
    typeaces.pack( )
    typeentre.pack( )
    facturation.pack(  )
    facturationentre.pack( )
    gestion.pack()
    messagegestion.pack()
    ges.mainloop()
    global num
    num=ligneentre.get()

def create_table(sheet ):
    List = []
    for value in sheet.iter_rows(min_row=2, max_row=200, min_col=2, max_col=8, values_only=True):
        List.append(value)
    return List
def find_from_num(num, List):

    for i in range (0, len(List)):
        if List[i][2] == num:
            return i
    return False
def parcours_ws_num(num, sheet):

    L = create_table(sheet)
    r = find_from_num(num, L)
    v=str(L[r])
    messagegestion['text']=v
    messagelogin.config(text="some text")


parcours_ws_num(num, sheetDR)
parcours_ws_num(num, sheetFM)
parcours_ws_num(num, sheetOr)
parcours_ws_num(num, sheetElhaouz)
parcours_ws_num(num, sheetTad)
parcours_ws_num(num, sheetSud)
parcours_ws_num(num, sheetRT)
parcours_ws_num(num, sheetSale)
parcours_ws_num(num, sheetCasaNord)
parcours_ws_num(num, sheetCasaSud)
parcours_ws_num(num, sheetNord)
parcours_ws_num(num, sheetmobile)



#---Window---#

#make window
window = tkinter.Tk()
#change title
window.title(" PAGE D'AUTHENTIFICATION")
#change size
window.geometry("270x210")
#change window colour
window.configure(bg="#39d972")
title1 = tkinter.Label(window, text="_____Authentification___", bg="#39d972")
usertitle = tkinter.Label(window, text="---Username---", bg="#39d972")
passtitle = tkinter.Label(window, text="---Password---", bg="#39d972")
messagelogin = tkinter.Label(window,text=None, bg="#39d972")
#text entry windows
user = tkinter.Entry(window)
passw = tkinter.Entry(window, show='*')
#buttons
go = tkinter.Button(window, text="Log in!",command=login, bg="#93ff00")
#pack widgets
title1.pack()
usertitle.pack()
user.pack()
passtitle.pack()
passw.pack()
go.pack()
messagelogin.pack()
#start window
window.mainloop()
